"""
Gisko MB monthly sales report generator.

Takes two Shopify CSV exports and produces a formatted Excel report
with per-payment-method amount splits for Agnė.

Inputs:
  orders_csv:       Shopify Orders export
                    (Admin → Orders → Export → "Export orders" → CSV)
  transactions_csv: Shopify Transaction histories export
                    (Admin → Orders → Export → "Export transaction histories" → CSV)

Output:
  xlsx with one row per order. Single-gateway orders have only
  "Apmokėjimo būdas (1)" + "Suma USD (1)" filled. Mixed-payment orders
  fill both gateway columns and are highlighted yellow. Refunded /
  cancelled orders are highlighted red.

Usage (CLI):
  python build_sales_report.py orders.csv transactions.csv output.xlsx
  python build_sales_report.py orders.csv transactions.csv output.xlsx --period "2026-05"

Web UI entry point:
  gisko_sales.run(orders_csv, transactions_csv, output_xlsx, period="")
"""

import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


GATEWAY_MAP = {
    "shopify_payments": "Shopify Payments",
    "paypal": "PayPal Express Checkout",
    "Airwallex Klarna": "Airwallex Klarna",
    "gift_card": "Gift Card",
    "Airwallex - Cards and Local Payment Methods": "Airwallex Cards",
    "shop_cash": "Shop Cash",
    "manual": "Manual",
}


def parse_dt(s):
    if pd.isna(s):
        return None
    try:
        return pd.to_datetime(s).tz_localize(None)
    except Exception:
        return None


def build_payment_splits(transactions_csv):
    """For each order, returns list of (gateway, total_captured) in chronological order.
    Only successful sale/capture transactions are counted."""
    tx = pd.read_csv(transactions_csv, low_memory=False)
    settled = tx[(tx["Kind"].isin(["sale", "capture"])) & (tx["Status"] == "success")].copy()
    settled["Gateway_Friendly"] = settled["Gateway"].map(lambda g: GATEWAY_MAP.get(g, g))

    def aggregate(group):
        g = group.sort_values("Created At")
        sums = {}
        order_seq = []
        for _, row in g.iterrows():
            gw = row["Gateway_Friendly"]
            if gw not in sums:
                sums[gw] = 0
                order_seq.append(gw)
            sums[gw] += row["Amount"]
        return [(gw, sums[gw]) for gw in order_seq]

    return settled.groupby("Name").apply(aggregate, include_groups=False).to_dict()


def build_orders_frame(orders_csv, splits):
    """Load orders CSV, filter to one row per order, attach payment splits."""
    df = pd.read_csv(orders_csv, low_memory=False)
    orders = df[df["Financial Status"].notna()].copy()
    orders["Created_dt"] = orders["Created at"].apply(parse_dt)
    orders["Paid_dt"] = orders["Paid at"].apply(parse_dt)
    orders["Cancelled_dt"] = orders["Cancelled at"].apply(parse_dt)
    orders = orders.sort_values("Created_dt").reset_index(drop=True)

    def get_split(name):
        parts = splits.get(name, [])
        m1 = parts[0][0] if len(parts) > 0 else ""
        a1 = parts[0][1] if len(parts) > 0 else 0
        m2 = parts[1][0] if len(parts) > 1 else ""
        a2 = parts[1][1] if len(parts) > 1 else 0
        return m1, a1, m2, a2

    orders[["M1", "A1", "M2", "A2"]] = orders["Name"].apply(lambda n: pd.Series(get_split(n)))
    return orders


def write_xlsx(orders, output_path, period_label):
    wb = Workbook()
    ws = wb.active
    # Excel sheet title max 31 chars
    sheet_title = f"Pardavimai {period_label}" if period_label else "Pardavimai"
    ws.title = sheet_title[:31]

    HEADER_FILL = PatternFill("solid", start_color="1F2937")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    TOTAL_FONT = Font(name="Arial", bold=True, size=10)
    TOTAL_FILL = PatternFill("solid", start_color="E5E7EB")
    SPLIT_FILL = PatternFill("solid", start_color="FEF3C7")
    RED_FILL = PatternFill("solid", start_color="FEE2E2")
    THIN = Side(border_style="thin", color="D1D5DB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")

    ws["A1"] = "GISKO MB - Pardavimų ataskaita"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:U1")

    ws["A2"] = (
        f"Laikotarpis: {period_label} · Operacijų: {len(orders)} · "
        f"Apmokėjimo būdai iš sėkmingų tranzakcijų (kind=sale/capture, status=success) · "
        f"Geltonai pažymėtos eilutės: mišrus apmokėjimas"
    )
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="6B7280")
    ws.merge_cells("A2:U2")

    headers = [
        "Eil. Nr.", "Dokumento Nr.", "Order ID", "Sukurta", "Apmokėta",
        "Pirkėjas", "El. paštas", "Šalis", "Miestas", "Valiuta",
        "Suma be PVM", "Pristatymas", "Mokesčiai / PVM", "Viso", "Grąžinta",
        "Apmokėjimo būdas (1)", "Suma USD (1)", "Apmokėjimo būdas (2)", "Suma USD (2)",
        "Statusas", "Atšaukta",
    ]
    HEADER_ROW = 4
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    for i in range(len(orders)):
        o = orders.iloc[i]
        r = HEADER_ROW + 1 + i
        has_split = bool(o["M2"])
        vals = [
            i + 1,
            o["Name"],
            int(o["Id"]) if pd.notna(o["Id"]) else None,
            o["Created_dt"],
            o["Paid_dt"],
            o["Billing Name"] if pd.notna(o["Billing Name"]) else "",
            o["Email"] if pd.notna(o["Email"]) else "",
            o["Billing Country"] if pd.notna(o["Billing Country"]) else "",
            o["Billing City"] if pd.notna(o["Billing City"]) else "",
            o["Currency"],
            float(o["Subtotal"]) if pd.notna(o["Subtotal"]) else 0,
            float(o["Shipping"]) if pd.notna(o["Shipping"]) else 0,
            float(o["Taxes"]) if pd.notna(o["Taxes"]) else 0,
            float(o["Total"]) if pd.notna(o["Total"]) else 0,
            float(o["Refunded Amount"]) if pd.notna(o["Refunded Amount"]) else 0,
            o["M1"],
            float(o["A1"]) if o["A1"] else 0,
            o["M2"] if has_split else "",
            float(o["A2"]) if has_split and o["A2"] else None,
            o["Financial Status"] if pd.notna(o["Financial Status"]) else "",
            o["Cancelled_dt"],
        ]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            if col_idx == 1:
                c.alignment = CENTER
            elif col_idx in (11, 12, 13, 14, 15, 17, 19):
                c.alignment = RIGHT
                c.number_format = "#,##0.00"
            elif col_idx in (4, 5, 21):
                c.alignment = CENTER
                c.number_format = "yyyy-mm-dd hh:mm"
            elif col_idx in (8, 10):
                c.alignment = CENTER
            else:
                c.alignment = LEFT
        if has_split:
            for col_idx in range(1, 22):
                ws.cell(row=r, column=col_idx).fill = SPLIT_FILL
        status = o["Financial Status"]
        cancelled = pd.notna(o["Cancelled_dt"])
        if status in ("refunded", "partially_refunded") or cancelled:
            for col_idx in range(1, 22):
                ws.cell(row=r, column=col_idx).fill = RED_FILL

    total_row = HEADER_ROW + 1 + len(orders)
    first_data = HEADER_ROW + 1
    last_data = total_row - 1
    label_cell = ws.cell(row=total_row, column=2, value="VISO")
    label_cell.font = TOTAL_FONT
    label_cell.fill = TOTAL_FILL
    label_cell.alignment = RIGHT
    label_cell.border = BORDER
    for col_idx in (1, 3, 4, 5, 6, 7, 8, 9, 10, 16, 18, 20, 21):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill = TOTAL_FILL
        c.border = BORDER
    for col_idx in (11, 12, 13, 14, 15, 17, 19):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}{first_data}:{col_letter}{last_data})")
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.alignment = RIGHT
        c.number_format = "#,##0.00"
        c.border = BORDER

    widths = {
        "A": 7, "B": 12, "C": 14, "D": 18, "E": 18, "F": 26, "G": 30, "H": 8,
        "I": 18, "J": 8, "K": 12, "L": 11, "M": 14, "N": 12, "O": 11,
        "P": 22, "Q": 13, "R": 22, "S": 13, "T": 18, "U": 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A{HEADER_ROW}:U{last_data}"

    wb.save(output_path)


def run(orders_csv, transactions_csv, output_xlsx, period=""):
    """Programmatic entry point used by the web UI.

    Returns the output path. Mirrors main() but without argparse / console I/O.
    """
    splits = build_payment_splits(transactions_csv)
    orders = build_orders_frame(orders_csv, splits)
    write_xlsx(orders, output_xlsx, period)
    return output_xlsx


def main():
    parser = argparse.ArgumentParser(description="Build Gisko MB monthly sales report")
    parser.add_argument("orders_csv", help="Shopify Orders export CSV")
    parser.add_argument("transactions_csv", help="Shopify Transaction histories export CSV")
    parser.add_argument("output_xlsx", help="Output Excel file path")
    parser.add_argument("--period", default="", help="Period label, e.g. '2026-05-01 to 2026-05-31'")
    args = parser.parse_args()

    splits = build_payment_splits(args.transactions_csv)
    orders = build_orders_frame(args.orders_csv, splits)
    write_xlsx(orders, args.output_xlsx, args.period)

    # Quick summary printed to console
    print(f"Saved: {args.output_xlsx}")
    print(f"Operations: {len(orders)}")
    print(f"Gross revenue: ${orders['Total'].sum():,.2f}")
    print(f"Refunded: ${orders['Refunded Amount'].sum():,.2f}")
    mixed = (orders["M2"] != "").sum()
    print(f"Mixed-payment orders: {mixed}")


if __name__ == "__main__":
    main()
